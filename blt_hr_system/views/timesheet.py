from django.urls import reverse
from django.template import loader, RequestContext
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, render, redirect
from .. import models
from .. import forms
from django.urls import reverse_lazy
from django.views.generic.edit import CreateView
from django.db import transaction
from django.contrib import messages
from django.contrib.auth.models import User
from django.template.loader import render_to_string
from django.core.mail import EmailMessage
import datetime
import datedelta
from django.db.models import Q
from collections import defaultdict
from django.db import connection
import pandas as pd
from datetime import datetime, timedelta
from django.forms import modelformset_factory
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import csv

# -----------------------------------------------------------------
# GENERAL/HELPER FUNCTIONS
# -----------------------------------------------------------------

ts_dict = {
    "1": ["Jul 01 2019", "Jul 14 2019"],
    "2": ["Jul 15 2019", "Jul 28 2019"],
    "3": ["Jul 29 2019", "Aug 11 2019"],
    "4": ["Aug 12 2019", "Aug 25 2019"],
    "5": ["Aug 26 2019", "Sep 08 2019"],
    "6": ["Sep 09 2019", "Sep 22 2019"],
    "7": ["Sep 23 2019", "Oct 06 2019"],
    "8": ["Oct 07 2019", "Oct 20 2019"],
    "9": ["Oct 21 2019", "Nov 03 2019"],
    "10": ["Nov 04 2019", "Nov 17 2019"],
    "11": ["Nov 18 2019", "Dec 01 2019"],
    "12": ["Dec 02 2019", "Dec 15 2019"],
    "13": ["Dec 16 2019", "Dec 29 2019"],
    "14": ["Dec 30 2019", "Jan 12 2020"],
    "15": ["Jan 13 2020", "Jan 26 2020"],
    "16": ["Jan 27 2020", "Feb 09 2020"],
    "17": ["Feb 10 2020", "Feb 23 2020"],
    "18": ["Feb 24 2020", "Mar 08 2020"],
    "19": ["Mar 09 2020", "Mar 22 2020"],
    "20": ["Mar 23 2020", "Apr 05 2020"],
    "21": ["Apr 06 2020", "Apr 19 2020"],
    "22": ["Apr 20 2020", "May 03 2020"],
    "23": ["May 04 2020", "May 17 2020"],
    "24": ["May 18 2020", "May 31 2020"],
    "25": ["Jun 01 2020", "Jun 14 2020"],
    "26": ["Jun 15 2020", "Jun 28 2020"],
    "27": ["Jun 29 2020", "Jul 12 2020"],
    "28": ["Jul 13 2020", "Jul 26 2020"],
    "29": ["Jul 27 2020", "Aug 09 2020"],
    "30": ["Aug 10 2020", "Aug 23 2020"],
    "31": ["Aug 24 2020", "Sep 06 2020"],
    "32": ["Sep 07 2020", "Sep 20 2020"],
    "33": ["Sep 21 2020", "Oct 04 2020"],
    "34": ["Oct 05 2020", "Oct 18 2020"],
    "35": ["Oct 19 2020", "Nov 01 2020"],
    "36": ["Nov 02 2020", "Nov 15 2020"],
    "37": ["Nov 16 2020", "Nov 29 2020"],
    "38": ["Nov 30 2020", "Dec 13 2020"],
    "39": ["Dec 14 2020", "Dec 27 2020"],
    "40": ["Dec 28 2020", "Jan 10 2021"],
    "41": ["Jan 11 2021", "Jan 24 2021"],
    "42": ["Jan 25 2021", "Feb 07 2021"],
    "43": ["Feb 08 2021", "Feb 21 2021"],
    "44": ["Feb 22 2021", "Mar 07 2021"],
    "45": ["Mar 08 2021", "Mar 21 2021"],
    "46": ["Mar 22 2021", "Apr 04 2021"],
    "47": ["Apr 05 2021", "Apr 18 2021"],
    "48": ["Apr 19 2021", "May 02 2021"],
    "49": ["May 03 2021", "May 16 2021"],
    "50": ["May 17 2021", "May 30 2021"],
    "51": ["May 31 2021", "Jun 13 2021"],
    "52": ["Jun 14 2021", "Jun 27 2021"],
    "53": ["Jun 28 2021", "Jul 11 2021"],
    "54": ["Jul 12 2021", "Jul 25 2021"],
    "55": ["Jul 26 2021", "Aug 08 2021"],
    "56": ["Aug 09 2021", "Aug 22 2021"],
    "57": ["Aug 23 2021", "Sep 05 2021"],
    "58": ["Sep 06 2021", "Sep 19 2021"],
    "59": ["Sep 20 2021", "Oct 03 2021"],
    "60": ["Oct 04 2021", "Oct 17 2021"],
    "61": ["Oct 18 2021", "Oct 31 2021"],
    "62": ["Nov 01 2021", "Nov 14 2021"],
    "63": ["Nov 15 2021", "Nov 28 2021"],
    "64": ["Nov 29 2021", "Dec 12 2021"],
    "65": ["Dec 13 2021", "Dec 26 2021"]
}

def export_ts(ts_period):
    category_df = pd.DataFrame(list(models.hourly_timesheet.objects.all()
        .filter(ts_period=ts_period, is_finalized=True) \
        .order_by('ts_period', 'sage_job__job_id') \
        .values('sage_job__job_id', 'sage_job__job_desc', 'hours', 'description',
                'start_date', 'end_date', 'ts_period',
                'employee_id__first_name', 'employee_id__last_name', 'employee_id')))
    if len(category_df.index) < 1:
        category_df = pd.DataFrame(columns=['sage_job__job_id', 'sage_job__job_desc', 
                'hours', 'description',
                'start_date', 'end_date', 'ts_period',
                'employee_id__first_name', 'employee_id__last_name', 'employee_id'])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-timesheet.xlsx'.format(
        date=(ts_dict[str(ts_period)][0] + "-" + ts_dict[str(ts_period)][1]),
    )
    workbook = Workbook()
    # Delete the default worksheet
    workbook.remove(workbook.active)
    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    # Define the column titles and widths
    columns = [
        ('Sage Job ID', 15),
        ('Sage Job Description', 40),
        ('Hours Worked', 15),
        ('Work Description', 50),
    ]
    df_tmp = category_df[['employee_id__first_name', 'employee_id__last_name', 'employee_id']]
    df_tmp = df_tmp.sort_values(['employee_id__first_name', 'employee_id__last_name'])
    df_tmp = df_tmp.drop_duplicates()
    for index, row in df_tmp.iterrows():
        worksheet = workbook.create_sheet(
            title=row['employee_id__first_name'] + ' ' + row['employee_id__last_name'],
            index=index,
        )
        row_num = 1
        iter_emp = row['employee_id']

        # Assign values, styles, and formatting for each cell in the header
        for col_num, (column_title, column_width) in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.border = border_bottom
            cell.alignment = centered_alignment
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = column_width

        # Iterate through all movies of a category
        category_ts = category_df.query('employee_id == @iter_emp')
        for index, job in category_ts.iterrows():
            row_num += 1
            
            # Define data and formats for each cell in the row
            row = [
                (job['sage_job__job_id'], 'Normal'),
                (job['sage_job__job_desc'], 'Normal'),
                (job['hours'], 'Normal'),
                (job['description'], 'Normal'),
            ]

            # Assign values, styles, and formatting for each cell in the row
            for col_num, (cell_value, cell_format) in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.style = cell_format
                if cell_format == 'Currency':
                    cell.number_format = '#,##0.00 €'
                if col_num == 4:
                    cell.number_format = '[h]:mm;@'
                cell.alignment = wrapped_alignment

    workbook.save(response)
    return response

def export_emp(user_id):
    category_df = pd.DataFrame(list(models.hourly_timesheet.objects.all()
        .filter(employee_id=user_id, is_finalized=True) \
        .order_by('ts_period', 'sage_job__job_id') \
        .values('sage_job__job_id', 'sage_job__job_desc', 'hours', 'description',
                'start_date', 'end_date', 'ts_period')))
    if len(category_df.index) < 1:
        category_df = pd.DataFrame(columns=['sage_job__job_id', 
            'sage_job__job_desc', 'hours', 'description',
                'start_date', 'end_date', 'ts_period'])
    user_queryset = User.objects.all().filter(id=user_id) \
        .values('first_name', 'last_name')
    for u in user_queryset:
        first_name = u['first_name']
        last_name = u['last_name']
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={emp}-timesheet.xlsx'.format(
        emp=(first_name + " " + last_name),
    )
    workbook = Workbook()
    # Delete the default worksheet
    workbook.remove(workbook.active)
    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    # Define the column titles and widths
    columns = [
        ('Sage Job ID', 15),
        ('Sage Job Description', 40),
        ('Hours Worked', 15),
        ('Work Description', 50),
    ]
    df_tmp = category_df[['start_date', 'end_date', 'ts_period']]
    df_tmp = df_tmp.drop_duplicates()
    for index, row in df_tmp.iterrows():
        worksheet = workbook.create_sheet(
            title=row['start_date'].strftime('%b %d %Y') + ' to ' + row['end_date'].strftime('%b %d %Y'),
            index=index,
        )
        row_num = 1
        iter_ts = row['ts_period']

        # Assign values, styles, and formatting for each cell in the header
        for col_num, (column_title, column_width) in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.border = border_bottom
            cell.alignment = centered_alignment
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = column_width

        # Iterate through all movies of a category
        category_ts = category_df.query('ts_period == @iter_ts')
        for index, job in category_ts.iterrows():
            row_num += 1
            
            # Define data and formats for each cell in the row
            row = [
                (job['sage_job__job_id'], 'Normal'),
                (job['sage_job__job_desc'], 'Normal'),
                (job['hours'], 'Normal'),
                (job['description'], 'Normal'),
            ]

            # Assign values, styles, and formatting for each cell in the row
            for col_num, (cell_value, cell_format) in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.style = cell_format
                if cell_format == 'Currency':
                    cell.number_format = '#,##0.00 €'
                if col_num == 4:
                    cell.number_format = '[h]:mm;@'
                cell.alignment = wrapped_alignment

    workbook.save(response)
    return response

def export_emp_ts(ts_period, user_id):
    category_queryset = models.hourly_timesheet.objects.all().filter(employee_id=user_id,
        ts_period=ts_period, is_finalized=True).order_by('sage_job__job_id') \
        .values('sage_job__job_id', 'sage_job__job_desc', 'hours', 'description')
    user_queryset = User.objects.all().filter(id=user_id) \
        .values('first_name', 'last_name')
    for u in user_queryset:
        first_name = u['first_name']
        last_name = u['last_name']
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-{emp}-timesheet.xlsx'.format(
        date=(ts_dict[str(ts_period)][0] + "-" + ts_dict[str(ts_period)][1]),
        emp=(first_name + " " + last_name),
    )
    workbook = Workbook()
    # Delete the default worksheet
    workbook.remove(workbook.active)
    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    # Define the column titles and widths
    columns = [
        ('Sage Job ID', 15),
        ('Sage Job Description', 40),
        ('Hours Worked', 15),
        ('Work Description', 50),
    ]
    worksheet = workbook.create_sheet(
        title=ts_dict[str(ts_period)][0] + " " + ts_dict[str(ts_period)][1],
        index=1,
    )
    row_num = 1
    # Assign values, styles, and formatting for each cell in the header
    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.border = border_bottom
        cell.alignment = centered_alignment
        # set column width
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width
    # Iterate through all movies of a category
    for category in category_queryset:
        row_num += 1
        # Define data and formats for each cell in the row
        row = [
            (category['sage_job__job_id'], 'Normal'),
            (category['sage_job__job_desc'], 'Normal'),
            (category['hours'], 'Normal'),
            (category['description'], 'Normal'),
        ]
        # Assign values, styles, and formatting for each cell in the row
        for col_num, (cell_value, cell_format) in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.style = cell_format
            if cell_format == 'Currency':
                cell.number_format = '#,##0.00 €'
            if col_num == 4:
                cell.number_format = '[h]:mm;@'
            cell.alignment = wrapped_alignment
    workbook.save(response)
    return response

# ------------------------------------------------------------------
# EMPLOYEE FUNCTIONS
# ------------------------------------------------------------------

def timesheet_export(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse('login'))
    user_id = request.user.id
    system_access = True
    system_user = models.Profile.objects.all().filter(user_id=user_id)
    if request.user.username != "system_admin" and not system_user.exists():
        system_access = False
    user_id = request.user.id
    if request.method == 'POST' and 'export' in request.POST: 
        ts_period = int(request.POST['ts_period'])
        return export_emp_ts(ts_period, user_id)
    if request.method == 'POST' and 'export_all' in request.POST:
        return export_emp(user_id)
    models.hourly_timesheet.objects.all().filter(employee_id=user_id, is_finalized=True)
    export_form = forms.timesheet_export()
    context = {'export_form':export_form,
                'system_access':system_access,
    }
    return render(request, 'timesheet/timesheet_export.html', context)

# Employee function to view all timesheets
def timesheet_home(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse('login'))
    user_id = request.user.id
    system_access = True
    system_user = models.Profile.objects.all().filter(user_id=user_id)
    if request.user.username != "system_admin" and not system_user.exists():
        system_access = False
    df = pd.DataFrame(list(models.hourly_timesheet.objects.all() \
            .order_by('start_date').values('employee_id', 'hours', 'start_date', 'end_date')))
    if len(df.index) < 1:
        df = pd.DataFrame(columns=['employee_id', 'hours', 'start_date', 'end_date'])
    user_id = request.user.id
    df = df.query('employee_id == @user_id')
    df_sum = df.groupby(['start_date'])['hours'].sum().reset_index()
    # create a list of the start dates where we require the employee to have a timesheet
    start_lst = [ ts_dict[k][0] for k in ts_dict ]
    now = datetime.now()
    required_lst = []
    for i in start_lst:
        if datetime.strptime(i, "%b %d %Y") <= now:
            add_date = datetime.strptime(i, "%b %d %Y").date()
            required_lst.append(add_date)
    df_to_print = pd.DataFrame(required_lst, columns = ['start_date'])
    df_to_print['end_date'] = df_to_print['start_date'] + timedelta(days=13)
    df_final_print = pd.merge(df_to_print, df_sum, how='left')
    df_final_print = df_final_print.fillna(0)
    df_final_print.hours = df_final_print.hours.astype(int)
    df_final_print = df_final_print.query('hours < 40')
    df_final_print_query = df_final_print.T.to_dict().values()
    df_sum = df_sum.query('hours >= 40')
    df_sum['end_date'] = df_sum['start_date'] + timedelta(days=13)
    df_sum = df_sum.T.to_dict().values()
    context = {'df_final_print_query':df_final_print_query,
                'df_sum':df_sum,
                'system_access':system_access,
    }
    return render(request, 'timesheet/timesheet_home.html', context)

def timesheet_hourly(request, pk):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse('login'))
    user_id = request.user.id
    system_access = True
    system_user = models.Profile.objects.all().filter(user_id=user_id)
    if request.user.username != "system_admin" and not system_user.exists():
        system_access = False
    start_date, end_date = ts_dict[pk]
    if request.method == 'POST':
        myModelFormset = modelformset_factory(models.hourly_timesheet, form=forms.hourly_ts
        , extra=0)
        items = models.hourly_timesheet.objects.filter(ts_period=pk, 
            employee_id=request.user.id).exclude(is_finalized=True).order_by('sage_job')
        formsetInstance = myModelFormset(queryset = items, data=request.POST)
        if formsetInstance.is_valid():
            # If there is already finalized data, don't let the form be re-submitted
            df = pd.DataFrame(list(models.hourly_timesheet.objects.all().values('id', 'employee_id', 
                'sage_job', 'is_finalized', 'hours', 'description', 'start_date', 'end_date', 'ts_period')))
            if len(df.index) < 1:
                df = pd.DataFrame(columns=['id', 'employee_id', 'sage_job', 'is_finalized', 'hours', 
                    'description', 'start_date', 'end_date', 'ts_period'])
            user_id = request.user.id
            df = df.query('employee_id == @user_id and ts_period == @pk')
            if len(df.index) > 0:
                if True in df['is_finalized'].tolist():
                    errors = ['Sorry, you have already uploaded your timesheet for this period.',
                      'Please contact HR if updates need to be made.']
                    myModelFormset = modelformset_factory(models.hourly_timesheet, form=forms.hourly_ts
                            , extra=0)
                    items = models.hourly_timesheet.objects.filter(ts_period=pk, 
                            employee_id=request.user.id).exclude(is_finalized=True).order_by('sage_job')
                    formsetInstance = myModelFormset(queryset = items)
                    context = {'start_date': start_date,
                                'end_date': end_date,
                                'formsetInstance': formsetInstance,
                                'errors':errors,
                                'system_access':system_access,
                                }
                    return render(request, 'timesheet/timesheet_hourly.html', context)
                else:
                    formsetInstance.save()
                    sum_hours = 0
                    for form_id in range(0, int(request.POST['form-TOTAL_FORMS'])):
                        hours_id = 'form-' + str(form_id) + '-hours'
                        tmp_hrs = int(request.POST[hours_id])
                        sum_hours += tmp_hrs
                    if sum_hours > 39:
                        items.update(is_finalized=True)
                    return HttpResponseRedirect(reverse('timesheet_home'))
    myModelFormset = modelformset_factory(models.hourly_timesheet, form=forms.hourly_ts
        , extra=0)
    items = models.hourly_timesheet.objects.filter(ts_period=pk, 
        employee_id=request.user.id).exclude(is_finalized=True).order_by('sage_job')
    formsetInstance = myModelFormset(queryset = items)
    errors = []
    context = {'start_date': start_date,
                'end_date': end_date,
                'formsetInstance': formsetInstance,
                'errors':errors,
                'system_access':system_access,
                }
    return render(request, 'timesheet/timesheet_hourly.html', context)

def val_job_form(user_id, ts_id, start_date, end_date, job_lst):
    job_lst = list(map(int, job_lst))
    df = pd.DataFrame(list(models.hourly_timesheet.objects.all().values('id', 'employee_id', 
        'sage_job', 'is_finalized', 'hours', 'description', 'start_date', 'end_date', 'ts_period')))
    if len(df.index) < 1:
        df = pd.DataFrame(columns=['id', 'employee_id', 'sage_job', 'is_finalized', 'hours', 
            'description', 'start_date', 'end_date', 'ts_period'])
    df = df.query('employee_id == @user_id and ts_period == @ts_id')
    if len(df.index) > 0:
        if True in df['is_finalized'].tolist():
            errors = ['Sorry, you have already uploaded your timesheet for this period.',
                      'After 40 hours have been inputted, only the Accounting Department can edit the timesheets.']
            page_type = 'timesheet/timesheet_center.html'
            return errors, page_type
        else:
            # first step is to delete any entries we don't want anymore
            to_delete_df = df[df['sage_job'].isin(job_lst) == False]
            delete_ids = set(to_delete_df['id'])
            for i in delete_ids:
                del_obj = models.hourly_timesheet.objects.filter(id=i).delete()
            # get the "to add" list by finding missing values from the db
            db_job_lst = df.sage_job.values.tolist()
            to_add_lst = list(set(job_lst) - set(db_job_lst))
            for i in to_add_lst:
                row = models.hourly_timesheet.objects.create(employee_id_id=user_id, sage_job_id=i,
                    is_finalized=False, hours=0, start_date=datetime.strptime(start_date, "%b %d %Y"),
                    end_date=datetime.strptime(end_date, "%b %d %Y"), ts_period=ts_id) 
    else:
        for i in job_lst:
            row = models.hourly_timesheet.objects.create(employee_id_id=user_id, sage_job_id=i,
                is_finalized=False, hours=0, start_date=datetime.strptime(start_date, "%b %d %Y"),
                end_date=datetime.strptime(end_date, "%b %d %Y"), ts_period=ts_id)
    page_type = 'timesheet_hourly'
    errors = []
    return errors, page_type
    

# Employee function to view complete a specific timesheet
def timesheet_center(request, pk):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse('login'))
    user_id = request.user.id
    system_access = True
    system_user = models.Profile.objects.all().filter(user_id=user_id)
    if request.user.username != "system_admin" and not system_user.exists():
        system_access = False
    select_jobs = forms.select_jobs()
    start_date, end_date = ts_dict[pk]
    if request.method == 'POST':
        job_lst = request.POST.getlist('Sage_jobs')
        if len(job_lst) < 1:
            errors = ['Please select at least one job for the timesheet.']
            context = {'select_jobs':select_jobs,
                        'start_date':start_date,
                        'end_date':end_date,
                        'errors':errors,
                        'system_access':system_access,
            }
            return render(request, 'timesheet/timesheet_center.html', context)
        else:
            errors, return_page = val_job_form(request.user.id, pk, start_date, end_date, job_lst)
            context = {'select_jobs':select_jobs,
                        'start_date':start_date,
                        'end_date':end_date,
                        'errors':errors,
                        'system_access':system_access,
            }
            if return_page == 'timesheet/timesheet_center.html':
                return render(request, return_page, context)
            else:
                return redirect('timesheet_hourly', pk=pk)
    errors = []
    context = {'select_jobs':select_jobs,
        'start_date':start_date,
        'end_date':end_date,
        'errors':errors,
        'system_access':system_access,
    }
    return render(request, 'timesheet/timesheet_center.html', context)

# Employee function to view all timesheets
def timesheet_selection(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse('login'))
    user_id = request.user.id
    system_access = True
    system_user = models.Profile.objects.all().filter(user_id=user_id)
    if request.user.username != "system_admin" and not system_user.exists():
        system_access = False
    if request.method == 'POST':
        user_id = request.user.id
        val = request.POST['ts_period']
        return redirect('timesheet_center', pk=val)
    else:
        timesheet_select = forms.timesheet_select()
        context = {'timesheet_select': timesheet_select,
                    'system_access':system_access,
        }
        return render(request, 'timesheet/timesheet_selection.html', context)

# ------------------------------------------------------------------
# ADMIN FUNCTIONS
# ------------------------------------------------------------------

# Admin function to upload job ID's
def jobs_upload(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse('login'))
    user_id = request.user.id
    system_access = True
    system_user = models.Profile.objects.all().filter(user_id=user_id)
    if request.user.username != "system_admin" and not system_user.exists():
        system_access = False
    all_user = models.Profile.objects.all().filter(user_id=user_id, access__access_level='All')
    ts_user = models.Profile.objects.all().filter(user_id=user_id, access__access_level='Timesheets')
    if request.user.username != "system_admin" and not all_user.exists() and not ts_user.exists():
        return HttpResponseRedirect(reverse('home'))
    if request.method == 'POST':
            form = forms.upload_jobs(request.POST, request.FILES)
            errors = []
            if form.is_valid():
                file = request.FILES['docfile']
                # first check to make sure a csv has been uploaded
                if file.name.split('.')[-1] != 'csv':
                    upload_jobs = forms.upload_jobs()
                    sage_jobs = models.sage_jobs.objects.all()
                    context = {'upload_jobs':upload_jobs,
                                'errors': errors,
                                'sage_jobs':sage_jobs}
                    errors.append('The file extension must end in .csv.')
                    return render(request, 'timesheet/jobs_upload.html', context)
                else:
                    # then check to make sure the csv is in the correct format
                    df = pd.read_csv(request.FILES['docfile'])
                    try:
                        # check a few conditions to make sure the date is ok
                        if df.count()[0] == df.count()[1]:
                            sage_jobs = pd.DataFrame(list(models.sage_jobs.objects.all().values('job_id', 'job_desc')))
                            if len(sage_jobs.index) < 1:
                                sage_jobs = pd.DataFrame(columns=['job_id', 'job_desc'])
                            job_id_lst = sage_jobs['job_id'].tolist()
                            df = df[~(df["Job"].isin(job_id_lst))]
                            for row in df.itertuples():
                                # TO DO: only add if the value does not exist in the db
                                row = models.sage_jobs.objects.create(job_id=row.Job, job_desc=row.Description)
                            return redirect('jobs_upload')
                        else:
                            upload_jobs = forms.upload_jobs()
                            sage_jobs = models.sage_jobs.objects.all()
                            context = {'upload_jobs':upload_jobs,
                                        'errors': errors,
                                        'sage_jobs':sage_jobs,
                                        'system_access':system_access,
                                        }
                            errors.append('There was an error in the uploaded csv. Please ensure:')
                            errors.append('You have a "Job" and "Description" column.')
                            errors.append('Both columns are the same length.')
                            return render(request, 'timesheet/jobs_upload.html', context)
                    except:
                        upload_jobs = forms.upload_jobs()
                        sage_jobs = models.sage_jobs.objects.all()
                        context = {'upload_jobs':upload_jobs,
                                    'errors': errors,
                                    'sage_jobs':sage_jobs,
                                    'system_access':system_access,
                                    }
                        errors.append('There was an error in the uploaded csv. Please ensure:')
                        errors.append('You have a "Job" and "Description" column.')
                        errors.append('Both columns are the same length.')
                        return render(request, 'timesheet/jobs_upload.html', context)
            else:
                upload_jobs = forms.upload_jobs()
                errors = []
                context = {'upload_jobs':upload_jobs,
                            'errors': errors,
                            'system_access':system_access,
                            }
                return render(request, 'timesheet/jobs_upload.html', context)
    upload_jobs = forms.upload_jobs()
    errors = []
    sage_jobs = models.sage_jobs.objects.all().order_by('job_id')
    context = {'upload_jobs':upload_jobs,
                'errors': errors,
                'sage_jobs':sage_jobs,
                'system_access':system_access,
                }
    return render(request, 'timesheet/jobs_upload.html', context)

# Admin function to edit the descriptions of job IDs
def job_upload_edit(request, pk):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse('login'))
    user_id = request.user.id
    system_access = True
    system_user = models.Profile.objects.all().filter(user_id=user_id)
    if request.user.username != "system_admin" and not system_user.exists():
        system_access = False
    all_user = models.Profile.objects.all().filter(user_id=user_id, access__access_level='All')
    ts_user = models.Profile.objects.all().filter(user_id=user_id, access__access_level='Timesheets')
    if request.user.username != "system_admin" and not all_user.exists() and not ts_user.exists():
        return HttpResponseRedirect(reverse('home'))
    if request.method == 'POST':
        sage_job = models.sage_jobs.objects.get(id=pk)
        sage_form = forms.sage_job_update(request.POST, instance=sage_job)
        if sage_form.is_valid():
            sage_form.save()
            return HttpResponseRedirect(reverse('jobs_upload'))
        else:
            messages.error(request, 'Please correct the error below.')
    sage_job = models.sage_jobs.objects.get(id=pk)
    sage_form = forms.sage_job_update(instance=sage_job)
    context = {'sage_form':sage_form,
                'system_access':system_access,
    }
    return render(request, 'timesheet/job_upload_edit.html', context)

# Admin function to employees that are and are not considered "office staff"
def admin_timesheet_home(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse('login'))
    user_id = request.user.id
    system_access = True
    system_user = models.Profile.objects.all().filter(user_id=user_id)
    if request.user.username != "system_admin" and not system_user.exists():
        system_access = False
    all_user = models.Profile.objects.all().filter(user_id=user_id, access__access_level='All')
    ts_user = models.Profile.objects.all().filter(user_id=user_id, access__access_level='Timesheets')
    if request.user.username != "system_admin" and not all_user.exists() and not ts_user.exists():
        return HttpResponseRedirect(reverse('home'))
    users = User.objects.all().order_by('-profile__office_staff', 'first_name', 'last_name')
    context = {'users' : users,
            'system_access':system_access,
    }
    return render(request, 'timesheet/admin_timesheet_home.html', context)

# Admin function to change the "office staff" for a particular employee
def admin_timesheet(request, pk):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse('login'))
    user_id = request.user.id
    system_access = True
    system_user = models.Profile.objects.all().filter(user_id=user_id)
    if request.user.username != "system_admin" and not system_user.exists():
        system_access = False
    all_user = models.Profile.objects.all().filter(user_id=user_id, access__access_level='All')
    ts_user = models.Profile.objects.all().filter(user_id=user_id, access__access_level='Timesheets')
    if request.user.username != "system_admin" and not all_user.exists() and not ts_user.exists():
        return HttpResponseRedirect(reverse('home'))
    if request.method == 'POST':
        user_account = User.objects.get(id=pk)
        office_form = forms.edit_office_staff(request.POST, instance=user_account.profile)
        office_form.save()
        messages.success(request, 'The employee office staff field was successfully updated!')
        return HttpResponseRedirect(reverse('admin_timesheet_home'))
    else:
        user_account = User.objects.get(id=pk)
        office_form = forms.edit_office_staff(instance=user_account.profile)
        context = {'user_account': user_account,
                    'office_form': office_form,
                    'system_access':system_access,
                    }
        return render(request, 'timesheet/admin_timesheet.html', context)

# Admin function to view the status of employee timesheets
def timesheet_status(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse('login'))
    user_id = request.user.id
    system_access = True
    system_user = models.Profile.objects.all().filter(user_id=user_id)
    if request.user.username != "system_admin" and not system_user.exists():
        system_access = False
    all_user = models.Profile.objects.all().filter(user_id=user_id, access__access_level='All')
    ts_user = models.Profile.objects.all().filter(user_id=user_id, access__access_level='Timesheets')
    if request.user.username != "system_admin" and not all_user.exists() and not ts_user.exists():
        return HttpResponseRedirect(reverse('home'))
    # get list of all office staff
    df_users = pd.DataFrame(list(models.Profile.objects.all() \
        .filter(office_staff=True) \
        .values('user', 'user__first_name','user__last_name')))
    if len(df_users.index) < 1:
        df_users = pd.DataFrame(columns=['user', 'user__first_name','user__last_name'])
    df_users = df_users.rename(columns={'user':'employee_id'})
    # get list of all completed timesheets
    df = pd.DataFrame(list(models.hourly_timesheet.objects.all() \
            .order_by('start_date').values('employee_id', 'hours', 'start_date')))
    if len(df.index) < 1:
        df = pd.DataFrame(columns=['start_date'])
    df_sum = df.groupby(['start_date', 'employee_id'])['hours'].sum().reset_index()
    # create a list of the start dates where we require the employee to have a timesheet
    start_lst = [ ts_dict[k][0] for k in ts_dict ]
    now = datetime.now()
    required_lst = []
    for i in start_lst:
        if datetime.strptime(i, "%b %d %Y") <= now:
            add_date = datetime.strptime(i, "%b %d %Y").date()
            required_lst.append(add_date)
    df_users_final = pd.DataFrame(columns=['employee_id', 'user__first_name','user__last_name'])
    for st_dt in required_lst: 
        df_users['start_date'] = st_dt
        df_users_final = df_users_final.append(df_users, sort=True)    
    df_users_final['end_date'] = df_users_final['start_date'] + timedelta(days=13)
    df_users_final = pd.merge(df_users_final, df_sum, on=['start_date', 'employee_id'], how='left')
    df_users_final = df_users_final.fillna(0)
    df_users_final.hours = df_users_final.hours.astype(int)
    df_users_final = df_users_final.query('hours < 40')
    df_users_final_qr = df_users_final.T.to_dict().values()
    if request.method == 'POST':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=missing_timesheets.csv'
        df_users_final = df_users_final.rename(columns={'user__first_name':'first_name',
                                                        'user__last_name': 'last_name'})
        df_users_final = df_users_final[['start_date', 'end_date', 'first_name', 'last_name', 'hours']]
        df_users_final.to_csv(path_or_buf=response, header=True, index=False, quoting=csv.QUOTE_NONNUMERIC,encoding='utf-8')
        return response
    context = {'df_users_final_qr':df_users_final_qr,
                'system_access':system_access,
    }
    return render(request, 'timesheet/timesheet_status.html', context)

# Admin function to export timesheets
def export_timesheets(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse('login'))
    user_id = request.user.id
    system_access = True
    system_user = models.Profile.objects.all().filter(user_id=user_id)
    if request.user.username != "system_admin" and not system_user.exists():
        system_access = False
    all_user = models.Profile.objects.all().filter(user_id=user_id, access__access_level='All')
    ts_user = models.Profile.objects.all().filter(user_id=user_id, access__access_level='Timesheets')
    if request.user.username != "system_admin" and not all_user.exists() and not ts_user.exists():
        return HttpResponseRedirect(reverse('home'))
    if request.method == 'POST' and 'export_emp_ts' in request.POST: 
        ts_period = int(request.POST['ts_period'])
        user_id = int(request.POST['employee_id'])
        return export_emp_ts(ts_period, user_id)
        errors = []
    if request.method == 'POST' and 'export_emp' in request.POST:
        user_id = int(request.POST['employee_id'])
        try:
            return export_emp(user_id)
        except:
            pass
    if request.method == 'POST' and 'export_ts' in request.POST:
        ts_period = int(request.POST['ts_period'])
        try:
            return export_ts(ts_period)
        except:
            pass
    export_ts_form = forms.timesheet_export()
    export_emp_form = forms.timesheet_emp()
    export_emp_ts_form = forms.timesheet_emp_ts()
    context = {'export_ts_form': export_ts_form,
            'export_emp_form':export_emp_form,
            'export_emp_ts_form': export_emp_ts_form,
            'system_access':system_access,
            }
    return render(request, 'timesheet/export_timesheets.html', context)

def edit_timesheet_home(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse('login'))
    user_id = request.user.id
    system_access = True
    system_user = models.Profile.objects.all().filter(user_id=user_id)
    if request.user.username != "system_admin" and not system_user.exists():
        system_access = False
    all_user = models.Profile.objects.all().filter(user_id=user_id, access__access_level='All')
    ts_user = models.Profile.objects.all().filter(user_id=user_id, access__access_level='Timesheets')
    if request.user.username != "system_admin" and not all_user.exists() and not ts_user.exists():
        return HttpResponseRedirect(reverse('home'))
    if request.method == 'POST':
        employee_id = request.POST['employee_id']
        ts_period = request.POST['ts_period']
        pk = employee_id + "-" + ts_period
        return redirect('edit_timesheet_jobs', pk=pk)
    ts_form = forms.timesheet_emp_ts()
    context = {'ts_form': ts_form,
                'system_access':system_access,
    }
    return render(request, 'timesheet/edit_timesheet_home.html', context)

def val_job_form_admin(user_id, ts_id, start_date, end_date, job_lst):
    job_lst = list(map(int, job_lst))
    df = pd.DataFrame(list(models.hourly_timesheet.objects.all().values('id', 'employee_id', 
        'sage_job', 'is_finalized', 'hours', 'description', 'start_date', 'end_date', 'ts_period')))
    if len(df.index) < 1:
        df = pd.DataFrame(columns=['id', 'employee_id', 'sage_job', 'is_finalized', 'hours', 
            'description', 'start_date', 'end_date', 'ts_period'])
    df = df.query('employee_id == @user_id and ts_period == @ts_id')
    if len(df.index) > 0:
        # first step is to delete any entries we don't want anymore
        to_delete_df = df[df['sage_job'].isin(job_lst) == False]
        delete_ids = set(to_delete_df['id'])
        for i in delete_ids:
            del_obj = models.hourly_timesheet.objects.filter(id=i).delete()
        # get the "to add" list by finding missing values from the db
        db_job_lst = df.sage_job.values.tolist()
        to_add_lst = list(set(job_lst) - set(db_job_lst))
        for i in to_add_lst:
            row = models.hourly_timesheet.objects.create(employee_id_id=user_id, sage_job_id=i,
                is_finalized=False, hours=0, start_date=datetime.strptime(start_date, "%b %d %Y"),
                end_date=datetime.strptime(end_date, "%b %d %Y"), ts_period=ts_id) 
    else:
        for i in job_lst:
            row = models.hourly_timesheet.objects.create(employee_id_id=user_id, sage_job_id=i,
                is_finalized=False, hours=0, start_date=datetime.strptime(start_date, "%b %d %Y"),
                end_date=datetime.strptime(end_date, "%b %d %Y"), ts_period=ts_id)
    errors = []
    return errors

def edit_timesheet_jobs(request, pk):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse('login'))
    user_id = request.user.id
    system_access = True
    system_user = models.Profile.objects.all().filter(user_id=user_id)
    if request.user.username != "system_admin" and not system_user.exists():
        system_access = False
    all_user = models.Profile.objects.all().filter(user_id=user_id, access__access_level='All')
    ts_user = models.Profile.objects.all().filter(user_id=user_id, access__access_level='Timesheets')
    if request.user.username != "system_admin" and not all_user.exists() and not ts_user.exists():
        return HttpResponseRedirect(reverse('home'))
    emp_id_str, ts_id_str = pk.split('-')
    emp_id = int(emp_id_str)
    ts_id = int(ts_id_str)
    select_jobs = forms.select_jobs()
    start_date, end_date = ts_dict[ts_id_str]
    user_query = User.objects.get(id=emp_id)
    first_name, last_name = user_query.first_name, user_query.last_name
    if request.method == 'POST':
        job_lst = request.POST.getlist('Sage_jobs')
        if len(job_lst) < 1:
            errors = ['Please select at least one job for the timesheet.']
            context = {'select_jobs':select_jobs,
                        'start_date':start_date,
                        'end_date':end_date,
                        'errors':errors,
                        'first_name': first_name,
                        'last_name': last_name,
                        'system_access':system_access,
            }
            return render(request, 'timesheet/edit_timesheet_jobs.html', context)
        else:
            errors = val_job_form_admin(emp_id_str, ts_id_str, start_date, end_date, job_lst)
            context = {'select_jobs':select_jobs,
                        'start_date':start_date,
                        'end_date':end_date,
                        'errors':errors,
                        'first_name': first_name,
                        'last_name': last_name,
                        'system_access':system_access,
            }
            return redirect('edit_timesheet_complete', pk=pk)
    errors = []
    context = {'select_jobs':select_jobs,
        'start_date':start_date,
        'end_date':end_date,
        'errors':errors,
        'first_name': first_name,
        'last_name': last_name,
        'system_access':system_access,
    }
    return render(request, 'timesheet/edit_timesheet_jobs.html', context)

def edit_timesheet_complete(request, pk):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse('login'))
    user_id = request.user.id
    system_access = True
    system_user = models.Profile.objects.all().filter(user_id=user_id)
    if request.user.username != "system_admin" and not system_user.exists():
        system_access = False
    all_user = models.Profile.objects.all().filter(user_id=user_id, access__access_level='All')
    ts_user = models.Profile.objects.all().filter(user_id=user_id, access__access_level='Timesheets')
    if request.user.username != "system_admin" and not all_user.exists() and not ts_user.exists():
        return HttpResponseRedirect(reverse('home'))
    emp_id_str, ts_id_str = pk.split('-')
    emp_id = int(emp_id_str)
    ts_id = int(ts_id_str)
    start_date, end_date = ts_dict[ts_id_str]
    user_query = User.objects.get(id=emp_id)
    first_name, last_name = user_query.first_name, user_query.last_name
    if request.method == 'POST':
        myModelFormset = modelformset_factory(models.hourly_timesheet, form=forms.hourly_ts
        , extra=0)
        items = models.hourly_timesheet.objects.filter(ts_period=ts_id, 
            employee_id=emp_id).order_by('sage_job')
        formsetInstance = myModelFormset(queryset = items, data=request.POST)
        print(formsetInstance)
        if formsetInstance.is_valid():
            print('valid')
            # If there is already finalized data, don't let the form be re-submitted
            df = pd.DataFrame(list(models.hourly_timesheet.objects.all().values('id', 'employee_id', 
                'sage_job', 'is_finalized', 'hours', 'description', 'start_date', 'end_date', 'ts_period')))
            if len(df.index) < 1:
                df = pd.DataFrame(columns=['id', 'employee_id', 'sage_job', 'is_finalized', 'hours', 
                    'description', 'start_date', 'end_date', 'ts_period'])
            user_id = request.user.id
            df = df.query('employee_id == @emp_id_str and ts_period == @ts_id_str')
            if len(df.index) > 0:
                formsetInstance.save()
                sum_hours = 0
                for form_id in range(0, int(request.POST['form-TOTAL_FORMS'])):
                    hours_id = 'form-' + str(form_id) + '-hours'
                    tmp_hrs = int(request.POST[hours_id])
                    sum_hours += tmp_hrs
                if sum_hours > 39:
                    items.update(is_finalized=True)
                    return HttpResponseRedirect(reverse('export_timesheets'))
                else:
                    items.update(is_finalized=False)
                    return HttpResponseRedirect(reverse('timesheet_status'))
    myModelFormset = modelformset_factory(models.hourly_timesheet, form=forms.hourly_ts
        , extra=0)
    items = models.hourly_timesheet.objects.filter(ts_period=ts_id, 
        employee_id=emp_id).order_by('sage_job')
    formsetInstance = myModelFormset(queryset = items)
    errors = []
    context = {'start_date': start_date,
                'end_date': end_date,
                'formsetInstance': formsetInstance,
                'errors':errors,
                'first_name':first_name,
                'last_name':last_name,
                'system_access':system_access,
                }
    return render(request, 'timesheet/edit_timesheet_complete.html', context)

